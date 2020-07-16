
from tkinter import *
from tkinter.filedialog import askopenfilename,askdirectory
from tkinter import messagebox
import time

import xlrd
from openpyxl import *
from openpyxl.styles.differential import DifferentialStyle
import openpyxl
from xlwt import  Workbook
import re
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, colors
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter.ttk import Progressbar

class job:
    def __init__(self,root):


        self.root = root


        self.depart()
        self.Liste_date = []
        self.Liste_issues = []
        self.Liste_review = []
        self.Liste_comments = []


    def depart(self):


        self.frame1 = Frame(self.root, width=600,bg='#FFFFFF')
        self.frame1.grid(row=0, column=0,ipady=10,ipadx=10)

        self.frame1.grid_columnconfigure(0,weight=2)

        self.vfichier1 = StringVar()
        self.vfichier2 = StringVar()
        self.vproject = StringVar()
        self.vfichier1.set('')
        self.vfichier2.set('')
        self.vproject.set('')
        self.chemin=''
        self.chemin1=''

        self.button1 = Button(self.frame1, text="Import weekly reporting", command=self.set_fichier1, width=50,height=2,bg='#66B239')
        self.button1.grid(row=0, column=0,pady=5)


        self.button2 = Button(self.frame1, text='Import SSR dashboard', command=self.set_fichier2, width=50,height=2,bg='#66B239')
        self.button2.grid(row=2, column=0, pady=5)

        self.vname_g=StringVar()
        self.vfilter_name=StringVar()
        self.vremarque=StringVar()
        self.vremarque.set("")
        self.vfilter_name.set('')


        self.bpetit_tab = Button(self.frame1, text='Import part list',command=self.set_filter, width=50,height=2,bg='#66B239')
        self.bpetit_tab.grid(row=4, column=0,pady=5)

        self.bProject = Button(self.frame1, text='Import Hard points list', command=self.set_project, width=50, height=2,
                                 bg='#66B239')
        self.bProject.grid(row=3, column=0, pady=5)

        self.bpetit_emp= Button(self.frame1, text="Generate report", command=self.set_emplacement, width=50, height=2,bg='#66B239')
        self.bpetit_emp.grid(row=5, column=0, pady=5)



        self.lremarque = Label(self.frame1, textvariable=self.vremarque, relief=GROOVE,
                               bg='#CAE21E', font=('arial', 10), border='2px', wraplength=550)
        self.progress_bar = Progressbar(self.frame1, orient='horizontal', length=286, mode='determinate')


    def part1(self,fichier1,fichier2,rapport_name,vproject,vfilter):
        self.progress_bar["maximum"] = 100

        self.bpetit_emp['bg'] = '#006738'
        self.progress_bar.grid(row=6, column=0, pady=2)
        self.progress_bar["value"] = 5
        root.update()
        # !/usr/bin/env python
        # coding: utf-8

        path = fichier1
        classeur = xlrd.open_workbook(path)
        self.progress_bar["value"] = 20
        root.update()
        nom_des_feuilles = classeur.sheet_names()
        #feuille = classeur.sheet_by_name(nom_des_feuilles[2])
        feuille = classeur.sheet_by_name("ECU Dashboard")

        #############################################################################################
        #############################################################################################
        def data_frame12(colonne1, colonne2, colonne3, colonne4, colonne5, colonne6):
            data = pd.DataFrame({"Part": (colonne1), "SSR Decision Trend": (colonne2), "Update_date": (colonne3),
                                 "Issues/Concerns": (colonne4), "Review/SSR_Status": (colonne5),
                                 "Expected Action": (colonne6)})
            return data
        dff=pd.read_excel(path,sheet_name="ECU Dashboard",skiprows=3)

        List_part=dff['ECU ID']
        liste_SSR_Status =dff["SSR status"]
        Sentence=dff["Review Plans & Commitments / Action plan/ Remarks"]


        term = "NEWDATE"

        term1 = "Issues/Concerns:"
        term2 = "Review/SSR Status:"
        term3 = "  "
        i=0
        for sentence in list(Sentence):
                    i+=1

                    sentence = str(sentence).replace('\n', 'sautligne')
                    sentence = sentence.replace('Review /SSR', 'Review/SSR')
                    sentence = sentence.replace('Review / SSR ', 'Review/SSR')
                    sentence = sentence.replace('Review/ SSR', 'Review/SSR')
                    sentence = sentence.replace('Review/SSR status', 'Review/SSR Status')
                    sentence = sentence.replace('Review/SSRstatus', 'Review/SSR Status')
                    sentence = sentence.replace('Review/SSRStatus', 'Review/SSR Status')
                    sentence = sentence.replace('Review/SSR Status :', 'Review/SSR Status:')

                    sentence = sentence.replace('Issues/ Concerns', 'Issues/Concerns')
                    sentence = sentence.replace('Issues /Concerns', 'Issues/Concerns')
                    sentence = sentence.replace('Issues / Concerns', 'Issues/Concerns')
                    sentence = sentence.replace('Issues/Concerns :', 'Issues/Concerns:')

                    list2 = re.findall("\d\d-\d\d-\d{4}", sentence)
                    for formatdate in list2:
                        sentence = sentence.replace(formatdate + " :", formatdate + ":")
                    try:
                        premieredate = list2[0]
                        list2 = [s for s in list2 if s != premieredate]
                        for formatdate in list2:
                            sentence = sentence.split(formatdate + ":")[0]
                            sentence = sentence.split(formatdate + "sautligne")[0]
                    except:
                        1
                    # on recupere le blocke le plus recent
                    block = sentence.split("NEWDATE")[0]

                    try:
                        if term1 in block and term2 in block and re.search('Issues/Concerns:(.*)Review/SSR Status:',
                                                                           block).group(1) != '' and re.search(
                            'Review/SSR Status:(.*)', block).group(1):
                            # on recupere la date (première occurence)
                            self.Liste_date.append(re.findall("\d\d-\d\d-\d{4}", block)[0])
                            # on recupere les Issues/Concerns (première occurence)
                            issue = block.split('Review/SSR Status:')[0]
                            issue = re.findall('Issues/Concerns:(.*)', issue)[0]
                            issue = issue.replace('sautlignesautligne', 'sautligne')
                            # on rajoute les retours à la ligne
                            try:
                                # print(re.search('sautligne(.*)', review).group(1))
                                if issue[0:9] == 'sautligne':
                                    issue = issue[9::]
                            except:
                                issue = issue

                            issue = issue.replace('sautligne', '\n')
                            self.Liste_issues.append(issue)
                            # on recupere les reviews (première occurence)

                            review = re.search('Review/SSR Status:(.*)', block).group(1)
                            self.List_date = re.findall("\d\d-\d\d-\d{4}", review)
                            for k in self.List_date:
                                review = review.split('sautligne' + k)[0]
                            review = review.replace('sautlignesautligne', 'sautligne')
                            # on rajoute les retours à la ligne
                            try:
                                # print(re.search('sautligne(.*)', review).group(1))
                                if review[0:9] == 'sautligne':
                                    review = review[9::]
                            except:
                                review = review

                            review = review.replace('sautligne', '\n')
                            self.Liste_review.append(review)
                            self.Liste_comments.append(term3)

                            # liste_comments.append(" {}".format(feuille.cell_value(i,64)))
                        else:
                            self.Liste_review.append(review)
                            self.Liste_comments.append(term3)
                            self.Liste_issues.append(issue)

                            self.Liste_date.append(".")
                    except:
                        self.Liste_date.append(".")
                        self.Liste_review.append(".")
                        self.Liste_comments.append(".")
                        self.Liste_issues.append(".")

        print(len(List_part),' ,',len(liste_SSR_Status),' ,',len(self.Liste_date),' ,',len(self.Liste_issues),' ,',len( self.Liste_review,),' ,',len( self.Liste_comments))
        ee=data_frame12(List_part, liste_SSR_Status, self.Liste_date, self.Liste_issues, self.Liste_review, self.Liste_comments)

        import numpy as np
        ea = ee.to_numpy()
        import matplotlib.pyplot as plt
        import numpy as np
        import datetime
        # premier tableau=================================================
        data = pd.read_excel(fichier2, "°Cover Page", skiprows=9, index_col=1)
        data = data.drop(['R_KPI_MILESTONE', 'Trend'], axis=1)
        e1 = data.iloc[0:4, 0:1]

        # deuxieme tableau===============================================
        data1 = pd.read_excel(fichier2, "°Cover Page", skiprows=43, index_col=2)
        data1 = data1.reset_index(drop=True)
        data1 = data1.drop(['Unnamed: 0', 'Unnamed: 1'], axis=1)
        e2 = data1.iloc[0:4, 0:11]
        self.progress_bar["value"] = 30
        root.update()
        time.sleep(0.5)
        # GRAPHE ==========================================================
        import matplotlib.pyplot as plt
        plt.rcParams.update({
            "pgf.texsystem": "pdflatex",
            "pgf.preamble": [
                r"\usepackage[utf8x]{inputenc}",
                r"\usepackage[T1]{fontenc}",
                r"\usepackage{cmbright}",
            ]
        })

        CWmax = e2['Unnamed: 3'][0].isocalendar()[1]
        x = []
        for i in range(CWmax, CWmax - 10, -1):
            x.append('CW' + str(i))

        self.progress_bar["value"] = 40
        root.update()
        y1 = e2.loc[1]
        y2 = e2.loc[2]
        y3 = e2.loc[3]
        plt.figure(figsize=(10, 5))
        plt.grid(True)
        plt.plot(x, y1, label='Coverage', lw=3)
        plt.plot(x, y2, label='Completness', lw=3)
        plt.plot(x, y3, label='Consistency', lw=3)
        self.progress_bar["value"] = 50
        time.sleep(1)
        root.update()
        plt.title('Milestone trend')
        plt.xlabel('Calendar Week')
        plt.ylabel('Kpi (%)')
        plt.legend()
        ax = plt.gca()
        ax.invert_xaxis()
        plt.savefig("fig.png")
        eb = e1.to_numpy()
        self.progress_bar["value"] = 60
        time.sleep(0.5)

        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import BarChart, Series, Reference

        path = vproject
        # data = pd.read_excel(path, skiprows=3, sheet_name=4)
        data = pd.read_excel(path, "Hard points", skiprows=3)
        id_list = list(data['ID'])
        veh_Project_list = list(data['Veh Project'])
        parts_list = list(data['Parts'])
        status_list = list(data['Status'])

        # Create a workbook and add a worksheet.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Report.xlsx'

        # Add a bold format to use to highlight cells.
        header_formatfont = Font(bold=True, )
        header_formattxt = Alignment(wrap_text=True)

        ## pour le petit tableau
        worksheet['A1'].value = 'KPI'
        worksheet['B1'].value = 'Completness'
        worksheet['C1'].value = 'Consistency'
        worksheet['B2'].value = 'Target = 100%'
        worksheet['C2'].value = 'Target = 80%'

        liste = ['A1', 'A2', 'B2', 'B1', 'C1', 'C2']
        for cell in liste:
            worksheet[cell].font = header_formatfont
            worksheet[cell].alignment = header_formattxt

        # data, workbook, and worksheet are the same as in the BarChart example
        tab = Table(displayName="Table1", ref="A1:C3")

        # I list out the 4 show-xyz options here for reference
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 70
        worksheet.column_dimensions['E'].width = 70
        worksheet.column_dimensions['F'].width = 40
        worksheet.column_dimensions['G'].width = 40
        self.progress_bar["value"] = 70
        time.sleep(1)
        root.update()

        # pour le grand tableau
        worksheet['A25'].value = 'PART'
        worksheet['B25'].value = 'SSR Decision Trend'
        worksheet['C25'].value = 'Update_date'
        worksheet['D25'].value = 'Issues/Concerns'
        worksheet['E25'].value = 'Review/SSR_Status'
        worksheet['F25'].value = 'Hard points and Risks IDs'
        worksheet['G25'].value = 'Expected Action'

        liste = ['A25', 'B25', 'C25', 'D25', 'E25', 'F25', 'G25']
        for cell in liste:
            worksheet[cell].font = header_formatfont
            worksheet[cell].alignment = header_formattxt

        # Petit tableau
        roww = 3
        coll = 0
        worksheet.cell(roww, 2).value = str(eb[1])
        worksheet.cell(roww, 3).value = str(eb[2])

        filename3 = vfilter
        wb3 = load_workbook(filename3)
        ws3 = wb3.worksheets[0]
        mr = ws3.max_row
        mc = ws3.max_column

        filter = []
        for i in range(1, mr + 1):
            c = ws3.cell(row=i, column=1)
            filter.append(c.value.upper())

        # Grand Tableau
        expenses = ea
        row = 26
        col = 1
        #expenses1 = ec

        col2 = 6

        for aa, bb, cc, dd, ff, gg  in (expenses):
            if str(aa).strip().upper() in filter:
                    worksheet.cell(row, col).value = aa
                    worksheet.cell(row, col).alignment = Alignment(wrapText=True, vertical='top')
                    worksheet.cell(row, col + 1).value = bb
                    worksheet.cell(row, col + 1).alignment = Alignment(wrapText=True, vertical='top')
                    worksheet.cell(row, col + 2).value = cc
                    worksheet.cell(row, col + 2).alignment = Alignment(vertical='top', wrapText=True)
                    worksheet.cell(row, col + 3).value = "'"+  str(dd).strip()
                    worksheet.cell(row, col + 3).alignment = Alignment(vertical='top', wrapText=True)
                    worksheet.cell(row, col + 4).value = "'"+ str(ff).strip()
                    worksheet.cell(row, col + 4).alignment = Alignment(vertical='top', wrapText=True)
                    worksheet.cell(row, col + 6).value = gg #str(gg).strip()
                    worksheet.cell(row, col + 6).alignment = Alignment(vertical='top', wrapText=True)

                    v_hp = ""
                    v_part = ""
                    v_final = ""

                    v_hp = ""
                    v_part = ""
                    v_final = ""
                    for i in range(len(id_list)):
                        v1 = str(veh_Project_list[i]).strip().upper() +"_" + str(parts_list[i]).strip().upper()

                        if v1 != v_part:

                            if str(aa).strip().upper() == v1.strip().upper():
                                if str(status_list[i]).strip().upper() == "OPEN":
                                    worksheet.cell(row, col2).value = str(id_list[i]) + '\n'
                                    worksheet.cell(row, col2).alignment = Alignment(wrapText=True, vertical='top')
                                    v_part = v1.strip()
                                    v_hp = ""
                                    v_final = id_list[i]
                        else:
                            if str(aa).strip().upper() == v1.strip().upper():
                                if str(status_list[i]).strip().upper() == "OPEN":
                                    v_hp += v_final + '\n'+  id_list[i]
                                    worksheet.cell(row, col2).value =v_hp
                                    worksheet.cell(row, col2).alignment = Alignment(wrapText=True, vertical='top')
                                    v_final = " "
                                    # v_final = aaa[0] + ' , '

                    row += 1
        piece_no_disponible = []
        piece_disponible = []
        self.progress_bar["value"] = 80
        time.sleep(1)
        root.update()
        for aa, bb, cc, dd, ff, gg in (expenses):
            piece_disponible.append(str(aa).upper().strip())
        for i in filter:
            if i not in piece_disponible:
                piece_no_disponible.append(i)
        # pour le message des piece non disponible
        l = ''
        for k in piece_no_disponible:
            if k != 'PART':
                l += ' , ' + str(k)

        li = 'The following parts ( ' + l + " ) are not available."

        if l != '':
            self.vremarque.set(li)

            self.lremarque['bg'] = '#FC4C4C'

        else:
            self.vremarque.set('Report created')
            self.lremarque['bg'] = '#CAE21E'

        #indice = len(expenses) + 25
        indice = len(filter) - len(piece_no_disponible) + 25
        ref = "A25:G" + str(indice)
        tab3 = Table(displayName="Table2", ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab3.tableStyleInfo = style
        worksheet.add_table(tab3)

        # Pour le graphe
        img = openpyxl.drawing.image.Image('fig.png')
        img.width = 750
        img.height = 370
        worksheet.add_image(img, 'A5')


        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_green = openpyxl.styles.colors.Color(rgb='0000FF00')
        my_orange = openpyxl.styles.colors.Color(rgb='00FFA500')
        # Couleur colonne B
        for i in range(26, len(expenses) + 26):

            if str(worksheet.cell(i, 2).value).strip() == 'Not Passed' or str(
                    worksheet.cell(i, 2).value).strip() == 'Not passed':
                worksheet.cell(i, 2).value = 'Not passed'
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_red)
            if str(worksheet.cell(i, 2).value).strip() == 'Conditionally':
                worksheet.cell(i, 2).value = worksheet.cell(i, 2).value.strip()
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_orange)
            if str(worksheet.cell(i, 2).value).strip() == 'Passed' or str(
                    worksheet.cell(i, 2).value).strip() == 'passed':
                worksheet.cell(i, 2).value = 'Passed'
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_green)
            """v = str(worksheet.cell(i, 2).value)
            if v.upper().strip() == 'NOT PASSED':
                worksheet.cell(i, 2).value = 'Not passed'
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_red)

            if str(worksheet.cell(i, 2).value).upper().strip() == 'CONDITIONALLY':
                worksheet.cell(i, 2).value = worksheet.cell(i, 2).value.strip()
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_orange)

            if str(worksheet.cell(i, 2).value).upper().strip() == 'PASSED':
                worksheet.cell(i, 2).value = 'Passed'
                worksheet.cell(i, 2).fill = PatternFill(patternType='solid', fgColor=my_green)"""

        workbook.save(rapport_name)
        workbook.close()
        self.progress_bar["value"] = 95
        time.sleep(2)
        root.update()
        self.progress_bar["value"] = 100
        time.sleep(2)

        root.update()

        self.progress_bar.grid_forget()
        root.update()

    def set_emplacement(self):
        import  time
        self.FILETYPES = [("text files", "*.xlsx")]
        self.chemin1=(askdirectory())
        date_now = time.strftime('%d%m%Y')
        self.chemin=self.chemin1+'/'+'F'+date_now+'.xlsx'
        if self.chemin1 !='':
            self.create_rapport(self.chemin)

    def set_fichier1(self):
        self.vremarque.set('')
        self.lremarque.grid_remove()
        self.FILETYPES = [("text files", "*.xlsx")]
        self.vfichier1.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier1.get()!='':
            self.button1['bg'] = '#006738'

    def set_fichier2(self):
        self.lremarque.grid_remove()
        self.vremarque.set('')
        self.FILETYPES = [("text files", "*.xlsx")]
        self.vfichier2.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier2.get()!='':
            self.button2['bg'] = '#006738'

    def set_project(self):
        self.vremarque.set('')
        self.lremarque.grid_remove()
        self.FILETYPES = [("text files", "*.xlsx")]
        self.vproject.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfichier1.get()!='':
            self.bProject['bg'] = '#006738'

    def set_filter(self):
        self.FILETYPES = [("text files", "*.xlsx")]
        self.vfilter_name.set(askopenfilename(filetypes=self.FILETYPES))
        if self.vfilter_name.get()!='':


           self.bpetit_tab['bg']='#006738'
    def create_rapport(self,name_r):

        if self.vfichier1.get()!='' and self.vfichier1.get()!='' and self.chemin1!='' and self.vfilter_name!='':

                self.part1(str(self.vfichier1.get()),self.vfichier2.get(),name_r,self.vproject.get(),self.vfilter_name.get())

                self.button2['bg'] = '#66B239'
                self.button1['bg'] = '#66B239'
                self.bpetit_tab['bg'] = '#66B239'
                self.bpetit_emp['bg'] = '#66B239'
                self.bProject['bg'] = '#66B239'
                self.lremarque.grid(row=7, column=0,sticky='news')

        else:
            messagebox.showerror('ERROR',"Import all files and select storage location")
if __name__ == '__main__':
     import tkinter as  tk

     root=Tk()
     job(root)
     root.mainloop()